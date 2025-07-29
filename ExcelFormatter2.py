def write_full_summary_list(summary_path, region_date_counts, all_dates, log_callback=None):
    """Create or update the summary_list.xlsx file for all regions and all dates in one pass."""
    import openpyxl
    from openpyxl.styles import Border, Side, Font, Alignment
    from openpyxl.utils import get_column_letter
    import calendar
    from datetime import datetime
    regions = [
        "NCR", "Region I", "Region II", "Region III", "Region IV", 
        "MIMAROPA", "Region V", "Region VI", "Region VII", "Region VIII", 
        "Region IX", "Region X", "Region XI", "Region XII", "CAR", 
        "Region XIII", "BARMM", "Unidentified"
    ]
    # Always create new workbook (overwrite)
    wb = openpyxl.Workbook()
    ws = wb.active
    # Determine year(s) from all_dates
    years = sorted({d.year for d in all_dates if hasattr(d, 'year')})
    if not years:
        title = "Daily Monitoring - Entry Count"
    elif len(years) == 1:
        title = f"{years[0]} Daily Monitoring - Entry Count"
    else:
        title = f"{years[0]}-{years[-1]} Daily Monitoring - Entry Count"
    ws['A1'] = title
    last_col_letter = get_column_letter(len(all_dates) + 1)
    # Unmerge any merged cells in the first row (shouldn't be any, but for safety)
    merged_ranges = list(ws.merged_cells.ranges)
    for rng in merged_ranges:
        if rng.min_row == 1:
            ws.unmerge_cells(str(rng))
    ws.merge_cells(f'A1:{last_col_letter}1')
    ws['A2'] = "REGIONS"
    # Date headers
    for i, date in enumerate(all_dates):
        col_letter = get_column_letter(i + 2)
        ws[f'{col_letter}2'] = format_date_header(date)
    # Region names
    for i, reg in enumerate(regions):
        ws[f'A{i+3}'] = reg
    # Apply formatting
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row in ws[f'A1:{last_col_letter}2']:
        for cell in row:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
    for row in range(3, len(regions) + 3):
        for col in range(1, len(all_dates) + 2):
            col_letter = get_column_letter(col)
            cell = ws[f'{col_letter}{row}']
            cell.border = thin_border
            if col == 1:
                cell.font = Font(bold=True)
            else:
                cell.alignment = Alignment(horizontal='center')
    # Fill in values
    for i, reg in enumerate(regions):
        reg_row = i + 3
        reg_counts = region_date_counts.get(reg, {})
        for j, date in enumerate(all_dates):
            col_letter = get_column_letter(j + 2)
            val = reg_counts.get(date, None)
            if val is not None and val != 0:
                ws[f'{col_letter}{reg_row}'].value = val
            else:
                ws[f'{col_letter}{reg_row}'].value = None
    wb.save(summary_path)
    if log_callback:
        log_callback(f"Full summary file written with {len(all_dates)} date columns and {len(regions)} regions: {summary_path}")
def compute_all_files(folder, summary_list_path=None, export_folder=None, log_callback=None):
    """
    1. Process all Excel files and sheets, writing per-file breakdowns to 'file_amounts.txt'.
    2. After all files, total all amounts per region and write to 'totaled_amounts.txt'.
    3. Generate summary_list.xlsx using totaled_amounts.txt as reference.
    """
    import json
    file_amounts_path = os.path.join(folder, 'file_amounts.txt')
    totaled_amounts_path = os.path.join(folder, 'totaled_amounts.txt')
    # Remove old files if exist
    for p in [file_amounts_path, totaled_amounts_path]:
        if os.path.exists(p):
            os.remove(p)

    excel_files = get_valid_excel_files(folder)
    region_date_amounts = defaultdict(lambda: defaultdict(int))  # region -> date -> amount
    all_dates = set()
    regions_set = set()

    with open(file_amounts_path, 'w', encoding='utf-8') as f_out:
        for file in excel_files:
            filename = os.path.basename(file)
            try:
                xl = pd.ExcelFile(file, engine='openpyxl')
                sheet_names = xl.sheet_names
            except Exception as e:
                if log_callback:
                    log_callback(f"  ERROR reading sheets: {e}")
                continue
            file_lines = [f"[{filename}]\n"]
            # Aggregate (date, region) -> count for this file
            file_agg = defaultdict(int)
            file_valid = False  # Track if any valid data for this file
            for sheet in sheet_names:
                try:
                    df = pd.read_excel(file, header=None, engine='openpyxl', skiprows=2, sheet_name=sheet)
                except Exception as e:
                    if log_callback:
                        log_callback(f"  ERROR reading sheet '{sheet}': {e}")
                    continue
                if df.shape[1] < 6 or df.shape[1] < 5:
                    continue
                df = df[df.iloc[:, 5].notna()]
                df = df[df.iloc[:, 5] != '-']
                df = df[df.iloc[:, 5].astype(str).str.strip().str.lower().isin(['local', 'imported'])]
                df.iloc[:, 0] = df.iloc[:, 0].apply(fix_date)
                df = df[df.iloc[:, 0].notna()]
                if len(df) == 0:
                    continue
                df['__region'] = df.iloc[:, 4].apply(lambda x: x.strip() if isinstance(x, str) else None)
                for idx, row in df.iterrows():
                    reg = row['__region']
                    dte = row.iloc[0]
                    amt = 1  # Each row is 1 entry (amount)
                    if isinstance(dte, pd.Timestamp):
                        dte = dte.date()
                    # Only use if dte is a date object and reg is a non-empty string
                    if not (isinstance(dte, date) and isinstance(reg, str) and reg.strip()):
                        if log_callback:
                            log_callback(f"    Skipped row {idx+2} in {filename} (invalid date or region: date={dte}, region={reg})")
                        continue
                    reg = reg.strip()
                    file_agg[(dte, reg)] += amt
                    file_valid = True
            # Only add to global aggregations if file_valid
            if file_valid:
                for (dte, reg), amt in file_agg.items():
                    region_date_amounts[reg][dte] += amt
                    all_dates.add(dte)
                    regions_set.add(reg)
            # Write one line per (date, region) for this file
            if log_callback:
                log_callback(f"  {filename}: {len(file_agg)} (date, region) pairs found.")
            for (dte, reg), amt in sorted(file_agg.items(), key=lambda x: (x[0][0], x[0][1])):
                file_lines.append(f"{dte} - {reg} - {amt}\n")
            if len(file_lines) > 1:
                f_out.writelines(file_lines)
                f_out.write("\n")
            else:
                if log_callback:
                    log_callback(f"  WARNING: No valid (date, region) pairs found for {filename}, nothing written to file_amounts.txt.")

    # Now total all amounts per region
    region_totals = defaultdict(int)
    with open(totaled_amounts_path, 'w', encoding='utf-8') as f_tot:
        for reg in sorted(regions_set):
            total = sum(region_date_amounts[reg].values())
            region_totals[reg] = total
            f_tot.write(f"{reg}: {total}\n")

    # Now generate summary_list.xlsx using region_date_amounts
    all_dates_sorted = sorted(all_dates)
    if not summary_list_path or not os.path.exists(summary_list_path):
        summary_path = os.path.join(export_folder if export_folder else folder, 'summary_list.xlsx')
    else:
        summary_path = summary_list_path

    # Write the full summary in one pass to avoid column duplication
    write_full_summary_list(summary_path, region_date_amounts, all_dates_sorted, log_callback)

    if log_callback:
        log_callback(f"file_amounts.txt and totaled_amounts.txt created. Summary list generated using totals.")

import os
import shutil
import glob
import pandas as pd
from datetime import datetime, timedelta, date
import tkinter as tk
from tkinter import filedialog, scrolledtext, messagebox
import threading
import openpyxl
from openpyxl.styles import Border, Side, Font, Alignment
import calendar
from collections import Counter, defaultdict

def clean_and_count_excel(file_path):
    """Clean and count entries from a single Excel file"""
    try:
        df = pd.read_excel(file_path, header=None, engine='openpyxl', skiprows=2)
        
        # Check if file has enough columns
        if df.shape[1] < 6:
            return 0
            
        # Start from third row (index 2) - already handled by skiprows=2
        df = df.copy()
        
        # Remove rows where column F (index 5) is blank or '-'
        df = df[df.iloc[:, 5].notna()]
        df = df[df.iloc[:, 5] != '-']
        
        # Keep only rows where column F is 'local' or 'imported' (case insensitive)
        df = df[df.iloc[:, 5].astype(str).str.strip().str.lower().isin(['local', 'imported'])]
        
        # Fix date format in column A (index 0)
        df.iloc[:, 0] = df.iloc[:, 0].apply(fix_date)
        
        # Remove rows with invalid dates
        df = df[df.iloc[:, 0].notna()]
        
        return len(df)
    except Exception as e:
        print(f"Error processing {file_path}: {e}")
        return 0

def fix_date(val):
    """Robust date parsing function"""
    if val is None or pd.isna(val):
        return None
        
    # Handle datetime objects first
    if isinstance(val, (pd.Timestamp, datetime)):
        return val.date()
    
    # Handle date objects
    if isinstance(val, date):
        return val
        

    # Handle string dates
    if isinstance(val, str):
        val = val.strip()
        if not val:
            return None

        # Handle date ranges: look for patterns like 'YYYY/MM/DD - YYYY/MM/DD' or 'YYYY/MM/DD-YY' or 'YYYY/MM/DD-YYYY/MM/DD' or 'YYYY/MM/DD-10'
        import re
        # Pattern 1: YYYY/MM/DD - YYYY/MM/DD or YYYY-MM-DD - YYYY-MM-DD
        range_match = re.match(r'^(\d{4}[/-]\d{2}[/-]\d{2})\s*[-–]\s*(\d{4}[/-]\d{2}[/-]\d{2})$', val)
        if range_match:
            end_str = range_match.group(2)
            try:
                return pd.to_datetime(end_str, errors='coerce').date()
            except:
                pass

        # Pattern 2: YYYY/MM/DD-YY or YYYY/MM/DD-10 (e.g., 2024/10/04-10)
        range_match2 = re.match(r'^(\d{4})[/-](\d{2})[/-](\d{2})\s*[-–]\s*(\d{2})$', val)
        if range_match2:
            year = int(range_match2.group(1))
            month = int(range_match2.group(2))
            start_day = int(range_match2.group(3))
            end_day = int(range_match2.group(4))
            # If end_day < start_day, assume next month
            if end_day < start_day:
                if month == 12:
                    year += 1
                    month = 1
                else:
                    month += 1
            try:
                return date(year, month, end_day)
            except:
                pass

        # Pattern 3: YYYY/MM/DD-YYYY/MM/DD (no spaces)
        range_match3 = re.match(r'^(\d{4}[/-]\d{2}[/-]\d{2})[-–](\d{4}[/-]\d{2}[/-]\d{2})$', val)
        if range_match3:
            end_str = range_match3.group(2)
            try:
                return pd.to_datetime(end_str, errors='coerce').date()
            except:
                pass

        # Pattern 4: YYYY/MM/DD-10 (no spaces)
        range_match4 = re.match(r'^(\d{4})[/-](\d{2})[/-](\d{2})[-–](\d{2})$', val)
        if range_match4:
            year = int(range_match4.group(1))
            month = int(range_match4.group(2))
            start_day = int(range_match4.group(3))
            end_day = int(range_match4.group(4))
            # If end_day < start_day, assume next month
            if end_day < start_day:
                if month == 12:
                    year += 1
                    month = 1
                else:
                    month += 1
            try:
                return date(year, month, end_day)
            except:
                pass

        # Try common date formats
        formats = [
            '%Y-%m-%d',      # 2024-03-07
            '%d/%m/%Y',      # 07/03/2024
            '%m/%d/%Y',      # 03/07/2024
            '%Y/%m/%d',      # 2024/03/07
            '%d-%m-%Y',      # 07-03-2024
            '%m-%d-%Y',      # 03-07-2024
            '%d.%m.%Y',      # 07.03.2024
            '%Y.%m.%d',      # 2024.03.07
            '%d %m %Y',      # 07 03 2024
            '%Y %m %d',      # 2024 03 07
        ]

        for fmt in formats:
            try:
                return datetime.strptime(val, fmt).date()
            except ValueError:
                continue

        # Try pandas parsing as last resort
        try:
            parsed = pd.to_datetime(val, errors='coerce', dayfirst=True)
            if not pd.isna(parsed):
                return parsed.date()
        except:
            pass

        return None
        
    # Handle numeric values (Excel serial dates)
    if isinstance(val, (int, float)):
        try:
            # Excel dates start from 1900-01-01 (but Excel thinks 1900 was a leap year)
            if val >= 1:
                # Convert Excel serial date to Python date
                if val < 60:  # Before March 1, 1900
                    base_date = datetime(1899, 12, 30)
                else:  # March 1, 1900 and after (account for fake leap day)
                    base_date = datetime(1899, 12, 31)
                    val -= 1
                
                result_date = base_date + timedelta(days=int(val))
                return result_date.date()
        except:
            pass
            
    return None

def format_date_header(date):
    """Format date as shown in the table (e.g., '7-Mar', '15-May')"""
    return f"{date.day}-{calendar.month_abbr[date.month]}"

def get_all_dates_from_folder(folder, log_callback=None):
    """Scan all Excel files in the folder to collect unique dates"""
    excel_files = get_valid_excel_files(folder)
    all_dates = set()
    
    for file in excel_files:
        try:
            df = pd.read_excel(file, header=None, engine='openpyxl', skiprows=2)

            # Check if enough columns
            if df.shape[1] < 6:
                if log_callback:
                    log_callback(f"File {os.path.basename(file)} does not have enough columns, skipping.")
                continue

            # Apply same filtering as main processing
            df = df[df.iloc[:, 5].notna()]
            df = df[df.iloc[:, 5] != '-']
            df = df[df.iloc[:, 5].astype(str).str.strip().str.lower().isin(['local', 'imported'])]

            # Fix dates
            df.iloc[:, 0] = df.iloc[:, 0].apply(fix_date)
            df = df[df.iloc[:, 0].notna()]

            # Add unique dates to our set
            file_dates = df.iloc[:, 0].unique()
            # Convert all pandas.Timestamp to datetime.date for comparison
            valid_dates = []
            for d in file_dates:
                if isinstance(d, pd.Timestamp):
                    valid_dates.append(d.date())
                elif isinstance(d, date):
                    valid_dates.append(d)
            all_dates.update(valid_dates)

        except Exception as e:
            if log_callback:
                log_callback(f"Error scanning dates from {os.path.basename(file)}: {e}")

    # Convert to sorted list
    sorted_dates = sorted(all_dates)

    if log_callback:
        log_callback(f"Found {len(sorted_dates)} unique dates across all files")
        if sorted_dates:
            log_callback(f"Date range: {sorted_dates[0]} to {sorted_dates[-1]}")

    return sorted_dates

def get_valid_excel_files(folder):
    """Get list of valid Excel files, excluding templates and temp files"""
    excel_files = glob.glob(os.path.join(folder, '*.xlsx'))
    
    # Exclude Template.xlsx, summary files, and Excel temp/lock files
    valid_files = []
    for f in excel_files:
        basename = os.path.basename(f).lower()
        if (not basename.startswith('template') and 
            not basename.startswith('summary') and 
            not basename.startswith('~$')):
            valid_files.append(f)
    
    return valid_files

def create_or_update_summary_list(summary_path, region, date_counts, all_dates, log_callback=None):
    """Create or update the summary_list.xlsx file with dynamic date columns"""
    
    regions = [
        "NCR", "Region I", "Region II", "Region III", "Region IV", 
        "MIMAROPA", "Region V", "Region VI", "Region VII", "Region VIII", 
        "Region IX", "Region X", "Region XI", "Region XII", "CAR", 
        "Region XIII", "BARMM", "Unidentified"
    ]
    
    # Check if file exists
    if os.path.exists(summary_path):
        # Load existing file
        wb = openpyxl.load_workbook(summary_path)
        ws = wb.active
        if log_callback:
            log_callback(f"Loading existing summary file: {summary_path}")

        # Unmerge any merged cells in the first row to avoid overlapping merges
        merged_ranges = list(ws.merged_cells.ranges)
        for rng in merged_ranges:
            if rng.min_row == 1:
                ws.unmerge_cells(str(rng))

        # Get existing date columns
        existing_dates = []
        for col in range(2, ws.max_column + 1):
            from openpyxl.utils import get_column_letter
            col_letter = get_column_letter(col)
            header_value = ws[f'{col_letter}2'].value
            if header_value:
                # Try to parse the existing header back to a date
                try:
                    day, month_abbr = header_value.split('-')
                    month_num = list(calendar.month_abbr).index(month_abbr)
                    # Use the year from our data
                    year = all_dates[0].year if all_dates else datetime.now().year
                    existing_date = datetime(year, month_num, int(day)).date()
                    existing_dates.append(existing_date)
                except:
                    if log_callback:
                        log_callback(f"Could not parse existing date header: {header_value}")
        # Do NOT add new columns for new dates here. Only update existing columns.
    else:
        # Create new workbook
        wb = openpyxl.Workbook()
        ws = wb.active

        # Set up headers
        from openpyxl.utils import get_column_letter
        ws['A1'] = "2025 Daily Monitoring - Entry Count"
        last_col_letter = get_column_letter(len(all_dates) + 1)
        # Unmerge any merged cells in the first row (shouldn't be any, but for safety)
        merged_ranges = list(ws.merged_cells.ranges)
        for rng in merged_ranges:
            if rng.min_row == 1:
                ws.unmerge_cells(str(rng))
        ws.merge_cells(f'A1:{last_col_letter}1')

        # Region header
        ws['A2'] = "REGIONS"

        # Date headers
        for i, date in enumerate(all_dates):
            col_letter = get_column_letter(i + 2)  # B, C, D, etc.
            ws[f'{col_letter}2'] = format_date_header(date)

        # Region names
        for i, reg in enumerate(regions):
            ws[f'A{i+3}'] = reg

        # Apply formatting
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Format headers
        for row in ws[f'A1:{last_col_letter}2']:
            for cell in row:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

        # Format region column and data cells
        for row in range(3, len(regions) + 3):
            for col in range(1, len(all_dates) + 2):
                col_letter = get_column_letter(col)
                cell = ws[f'{col_letter}{row}']
                cell.border = thin_border
                if col == 1:  # Region column
                    cell.font = Font(bold=True)
                else:
                    cell.alignment = Alignment(horizontal='center')

        if log_callback:
            log_callback(f"Created new summary file with {len(all_dates)} date columns: {summary_path}")
    
    # Update the cells with the counts
    from openpyxl.utils import get_column_letter

    # Find the correct row for the region
    region_row = None
    for row in range(3, ws.max_row + 1):
        if ws[f'A{row}'].value == region:
            region_row = row
            break

    if region_row is None:
        if log_callback:
            log_callback(f"Warning: Region '{region}' not found in summary file")
        wb.save(summary_path)
        return

    # Update counts for dates that have data
    for date, count in date_counts.items():
        # Find the correct column for this date
        date_col = None
        for col in range(2, ws.max_column + 1):
            col_letter = get_column_letter(col)
            header_date = ws[f'{col_letter}2'].value
            if header_date == format_date_header(date):
                date_col = col
                break

        if date_col is None:
            if log_callback:
                log_callback(f"Warning: Date column for '{format_date_header(date)}' not found in summary file. Skipping update for this date.")
            continue

        # Update the cell (ADD to existing value, don't replace)
        col_letter = get_column_letter(date_col)
        current_value = ws[f'{col_letter}{region_row}'].value
        if current_value is None:
            current_value = 0

        new_value = current_value + count
        ws[f'{col_letter}{region_row}'].value = new_value

        if log_callback:
            log_callback(f"Updated {region} for {format_date_header(date)}: {current_value} + {count} = {new_value}")

    # Save the file
    wb.save(summary_path)

def process_excels(folder, region, summary_list_path, log_callback=None, export_folder=None):
    """Main processing function"""
    # First, scan all files to get unique dates
    if log_callback:
        log_callback("=" * 50)
        log_callback("Starting Excel processing...")
        log_callback("Scanning all files to determine date columns...")
    

    # --- JSON tracking integration ---
    import json
    recorded_json_path = os.path.join(folder, 'recorded_files.json')
    recorded_files = set()
    # Load recorded files if exists
    if os.path.exists(recorded_json_path):
        try:
            with open(recorded_json_path, 'r', encoding='utf-8') as f:
                recorded_files = set(json.load(f))
        except Exception as e:
            if log_callback:
                log_callback(f"WARNING: Could not read recorded_files.json: {e}")
            recorded_files = set()
    else:
        # If file doesn't exist, create it
        with open(recorded_json_path, 'w', encoding='utf-8') as f:
            json.dump([], f)

    all_dates = get_all_dates_from_folder(folder, log_callback)

    if not all_dates:
        if log_callback:
            log_callback("ERROR: No valid dates found in any files!")
        return

    excel_files = get_valid_excel_files(folder)

    if not excel_files:
        if log_callback:
            log_callback("ERROR: No valid Excel files found!")
        return

    # Filter out already processed files
    files_to_process = [f for f in excel_files if os.path.basename(f) not in recorded_files]
    skipped_files = [f for f in excel_files if os.path.basename(f) in recorded_files]

    if log_callback:
        log_callback(f"Found {len(excel_files)} valid Excel files to process")
        if skipped_files:
            log_callback(f"Skipping {len(skipped_files)} files already processed (tracked in recorded_files.json):")
            for f in skipped_files:
                log_callback(f"  SKIPPED: {os.path.basename(f)}")

    # Prepare summary file path
    if not summary_list_path or not os.path.exists(summary_list_path):
        summary_path = os.path.join(export_folder if export_folder else folder, 'summary_list.xlsx')
        if log_callback:
            log_callback(f"Will create new summary file at: {summary_path}")
    else:
        summary_path = summary_list_path
        if log_callback:
            log_callback(f"Will update existing summary file: {summary_path}")

    # Region mapping for flexible matching
    region_map = {
        'NCR': ['NCR'],
        'Region I': ['REGION 1', 'REGION I', 'I', '1', 'REGION1', 'REGIONI'],
        'Region II': ['REGION 2', 'REGION II', 'II', '2', 'REGION2', 'REGIONII'],
        'Region III': ['REGION 3', 'REGION III', 'III', '3', 'REGION3', 'REGIONIII'],
        'Region IV': [
            'REGION 4', 'REGION IV', 'IV', '4', 'REGION4', 'REGIONIV',
            'REGION 4A', 'REGION IV-A', 'IV-A', '4A', 'IVA', 'REGION4A', 'REGIONIVA',
            'REGION 4B', 'REGION IV-B', 'IV-B', '4B', 'IVB', 'REGION4B', 'REGIONIVB',
            'REGIONIV-B', 'REGION IVB', 'REGION IV B', 'REGION 4 B', '4-B', 'IV B', 'IV-B', '4-B', 'IVB', '4 B',
        ],
        'MIMAROPA': [
            'MIMAROPA', 'REGION MIMAROPA', 'REGION-MIMAROPA', 'REGION_MIMAROPA',
            'REGIONMIMAROPA', 'M I M A R O P A', 'MIMAROPA REGION',
        ],
        'Region V': ['REGION 5', 'REGION V', 'V', '5', 'REGION5', 'REGIONV'],
        'Region VI': ['REGION 6', 'REGION VI', 'VI', '6', 'REGION6', 'REGIONVI'],
        'Region VII': ['REGION 7', 'REGION VII', 'VII', '7', 'REGION7', 'REGIONVII'],
        'Region VIII': ['REGION 8', 'REGION VIII', 'VIII', '8', 'REGION8', 'REGIONVIII'],
        'Region IX': ['REGION 9', 'REGION IX', 'IX', '9', 'REGION9', 'REGIONIX'],
        'Region X': ['REGION 10', 'REGION X', 'X', '10', 'REGION10', 'REGIONX'],
        'Region XI': ['REGION 11', 'REGION XI', 'XI', '11', 'REGION11', 'REGIONXI'],
        'Region XII': ['REGION 12', 'REGION XII', 'XII', '12', 'REGION12', 'REGIONXII'],
        'Region XIII': ['REGION 13', 'REGION XIII', 'XIII', '13', 'REGION13', 'REGIONXIII'],
        'CAR': ['CAR'],
        'BARMM': ['BARMM', 'Barmm'],
    }

    # Helper to normalize region name
    def normalize_region(val):
        # Handle NaN, None, blank
        if val is None:
            return 'Unidentified'
        # Convert float/int to string (e.g., 5.0 -> '5')
        if isinstance(val, float):
            if pd.isna(val):
                return 'Unidentified'
            # If float is integer-like, convert to int string
            if val.is_integer():
                val = str(int(val))
            else:
                val = str(val)
        elif isinstance(val, int):
            val = str(val)
        # Convert to string if not already
        if not isinstance(val, str):
            val = str(val)
        val = val.strip()
        if not val:
            return 'Unidentified'
        val_norm = val.upper().replace('-', '').replace(' ', '')
        for std, aliases in region_map.items():
            for alias in aliases:
                alias_norm = alias.strip().upper().replace('-', '').replace(' ', '')
                if val_norm == alias_norm:
                    return std
        # If not found, try to catch common patterns for 4B, IV-B, etc. (but NOT MIMAROPA)
        if val_norm in ['4B', 'IVB', 'IV-B', '4-B', 'IV B', '4 B']:
            return 'Region IV'
        # If not found, try to catch common patterns for MIMAROPA
        if val_norm in ['MIMAROPA', 'REGIONMIMAROPA', 'REGION-MIMAROPA', 'REGION_MIMAROPA', 'MIMAROPAREGION', 'MIMAROPAA']:
            return 'MIMAROPA'
        return 'Unidentified'

    finished_dir = os.path.join(folder, 'finished files')
    os.makedirs(finished_dir, exist_ok=True)



    # Process files one at a time, but only update summary at the end
    region_date_counts = defaultdict(lambda: defaultdict(int))
    total_count = 0
    results = []
    merged_df_list = []
    file_count = 0
    file_amounts_path = os.path.join(folder, 'file_amounts.txt')
    # Open file_amounts.txt for writing
    import time
    for file in files_to_process:
        file_count += 1
        filename = os.path.basename(file)
        try:
            if log_callback:
                log_callback(f"Processing file {file_count}/{len(files_to_process)}: {filename}")
            # Get all sheet names and ensure ExcelFile is closed after use
            try:
                with pd.ExcelFile(file, engine='openpyxl') as xl:
                    sheet_names = xl.sheet_names
            except Exception as e:
                if log_callback:
                    log_callback(f"  ERROR reading sheets: {e}")
                continue
            file_total = 0
            file_region_date_counts = defaultdict(lambda: defaultdict(int))
            file_merged_dfs = []
            valid_sheet_found = False
            file_agg = defaultdict(int)
            file_lines = [f"[{filename}]\n"]
            # Read all sheets into memory before moving file
            sheet_dfs = {}
            for sheet in sheet_names:
                try:
                    df = pd.read_excel(file, header=None, engine='openpyxl', skiprows=2, sheet_name=sheet)
                    sheet_dfs[sheet] = df.copy()
                except Exception as e:
                    if log_callback:
                        log_callback(f"  ERROR reading sheet '{sheet}': {e}")
                    continue
            for sheet, df in sheet_dfs.items():
                if df.shape[1] < 6:
                    if log_callback:
                        log_callback(f"  SKIPPED sheet '{sheet}': Not enough columns ({df.shape[1]} < 6)")
                    continue
                if df.shape[1] < 5:
                    if log_callback:
                        log_callback(f"  SKIPPED sheet '{sheet}': Column E (region) missing")
                    continue
                original_rows = len(df)
                df = df[df.iloc[:, 5].notna()]
                df = df[df.iloc[:, 5] != '-']
                df = df[df.iloc[:, 5].astype(str).str.strip().str.lower().isin(['local', 'imported'])]
                # Try to fix dates in column A
                df.iloc[:, 0] = df.iloc[:, 0].apply(fix_date)
                # If all dates are missing, try to extract end date from sheet name
                if df.iloc[:, 0].isna().all():
                    import re
                    sheet_name = str(sheet)
                    m = re.search(r'(\w+)\.?\s*(\d+)[-–](\w+\.?\s*)?(\d+)', sheet_name)
                    if m:
                        month1 = m.group(1)
                        day1 = m.group(2)
                        month2 = m.group(3)
                        day2 = m.group(4)
                        month_str = month2.strip() if month2 else month1.strip()
                        try:
                            from datetime import datetime as dt
                            try:
                                month_num = dt.strptime(month_str[:3], '%b').month
                            except:
                                month_num = dt.strptime(month_str, '%B').month
                            year = datetime.now().year
                            end_date = date(year, month_num, int(day2))
                            df.iloc[:, 0] = end_date
                            if log_callback:
                                log_callback(f"    All dates missing, using end date {end_date} from sheet name '{sheet_name}'")
                        except Exception as e:
                            if log_callback:
                                log_callback(f"    Could not parse end date from sheet name '{sheet_name}': {e}")
                df = df[df.iloc[:, 0].notna()]
                final_rows = len(df)
                if final_rows == 0:
                    if log_callback:
                        log_callback(f"  SKIPPED sheet '{sheet}': No valid data after filtering (started with {original_rows} rows)")
                    continue
                valid_sheet_found = True
                for idx, row in df.iterrows():
                    original_region = row.iloc[4] if len(row) > 4 else None
                    reg = normalize_region(original_region)
                    dte = row.iloc[0]
                    amt = 1
                    if log_callback:
                        log_callback(f"    Row {idx+2}: original region='{original_region}' -> assigned region='{reg}'")
                    if isinstance(dte, pd.Timestamp):
                        dte = dte.date()
                    if not (isinstance(dte, date) and isinstance(reg, str) and reg.strip()):
                        if log_callback:
                            log_callback(f"    Skipped row (invalid date or region: date={dte}, region={reg})")
                        continue
                    reg = reg.strip()
                    file_region_date_counts[reg][dte] += 1
                    region_date_counts[reg][dte] += 1
                    file_total += 1
                    file_agg[(dte, reg)] += 1
                file_merged_dfs.append(df)
            if not valid_sheet_found:
                if log_callback:
                    log_callback(f"  SKIPPED: No valid data in any sheet")
                continue
            total_count += file_total
            results.append(f"{filename} - {file_total}")
            merged_df_list.extend(file_merged_dfs)
            if log_callback:
                log_callback(f"  {filename}: {len(file_agg)} (date, region) pairs found.")
            for (dte, reg), amt in sorted(file_agg.items(), key=lambda x: (x[0][0], x[0][1])):
                file_lines.append(f"{dte} - {reg} - {amt}\n")
            if len(file_lines) > 1:
                with open(file_amounts_path, 'a', encoding='utf-8') as f_append:
                    f_append.writelines(file_lines)
                    f_append.write("\n")
            else:
                if log_callback:
                    log_callback(f"  WARNING: No valid (date, region) pairs found for {filename}, nothing written to file_amounts.txt.")
            recorded_files.add(filename)
            dest = os.path.join(finished_dir, filename)
            base, ext = os.path.splitext(filename)
            counter = 1
            while os.path.exists(dest):
                dest = os.path.join(finished_dir, f"{base}_moved{counter}{ext}")
                counter += 1
            # Try to move file, retry if locked (WinError 32)
            move_success = False
            for attempt in range(3):
                try:
                    shutil.move(file, dest)
                    move_success = True
                    break
                except Exception as move_e:
                    if 'WinError 32' in str(move_e) or 'being used by another process' in str(move_e):
                        time.sleep(0.5)
                    else:
                        break
            if move_success:
                if log_callback:
                    date_breakdown = []
                    for reg in sorted(file_region_date_counts.keys()):
                        date_keys = [d.date() if isinstance(d, pd.Timestamp) else d for d in file_region_date_counts[reg].keys()]
                        for d in sorted(date_keys):
                            date_breakdown.append(f"{reg}:{format_date_header(d)}({file_region_date_counts[reg][d]})")
                    date_summary = ", ".join(date_breakdown)
                    log_callback(f"  SUCCESS: {file_total} entries | Per region/date: {date_summary}")
                    log_callback(f"  Filtered: {original_rows} → {final_rows} rows")
            else:
                if log_callback:
                    log_callback(f"  ERROR moving file to finished files after retries.")
            try:
                with open(recorded_json_path, 'w', encoding='utf-8') as f:
                    json.dump(sorted(list(recorded_files)), f, indent=2)
                if log_callback:
                    log_callback(f"Updated recorded_files.json with {len(recorded_files)} processed files.")
            except Exception as e:
                if log_callback:
                    log_callback(f"WARNING: Could not update recorded_files.json: {e}")
        except Exception as e:
            if log_callback:
                log_callback(f"  ERROR processing {filename}: {e}")
            continue

    # After all files are processed, update the summary list ONCE
    try:
        if log_callback:
            log_callback("\nUpdating summary file for all regions...")
        # Collect the full set of all unique dates across all regions
        all_dates_set = set()
        for reg in region_date_counts:
            for d in region_date_counts[reg]:
                if isinstance(d, pd.Timestamp):
                    d = d.date()
                all_dates_set.add(d)
        all_dates_sorted = sorted(all_dates_set)
        # Write the full summary in one pass to avoid column duplication
        write_full_summary_list(summary_path, region_date_counts, all_dates_sorted, log_callback)
        if log_callback:
            log_callback("Summary file updated for all regions.")
    except Exception as e:
        if log_callback:
            log_callback(f"ERROR updating summary list at end: {e}")

    # Final summary of processing
    if log_callback:
        log_callback("=" * 50)
        log_callback(f"PROCESSING SUMMARY:")
        log_callback(f"Files processed successfully: {len(results)}")
        log_callback(f"Total entries found: {total_count}")
        log_callback(f"Date breakdown:")
        for reg in sorted(region_date_counts.keys()):
            # Convert all keys to datetime.date for sorting
            date_keys = [d.date() if isinstance(d, pd.Timestamp) else d for d in region_date_counts[reg].keys()]
            for dte in sorted(date_keys):
                log_callback(f"  {reg} - {format_date_header(dte)}: {region_date_counts[reg][dte]} entries")

    # Create merged file if template exists (after all files)
    template_path = os.path.join(folder, 'Template.xlsx')
    if merged_df_list and os.path.exists(template_path):
        try:
            if log_callback:
                log_callback("Creating merged file...")
            merged_df = pd.concat(merged_df_list, ignore_index=True)
            values_to_paste = merged_df.values.tolist()
            wb_template = openpyxl.load_workbook(template_path)
            ws_template = wb_template.active
            # Paste values starting from row 3
            for i, row_values in enumerate(values_to_paste, start=3):
                for j, value in enumerate(row_values, start=1):
                    if j <= ws_template.max_column:
                        ws_template.cell(row=i, column=j, value=value)
            folder_name = os.path.basename(folder)
            merged_path = os.path.join(folder, f'{folder_name}_merged.xlsx')
            wb_template.save(merged_path)
            if log_callback:
                log_callback(f"Merged file created: {merged_path}")
        except Exception as e:
            if log_callback:
                log_callback(f"ERROR creating merged file: {e}")

    if log_callback:
        log_callback("=" * 50)
        log_callback("PROCESSING COMPLETE!")
        log_callback(f"Final total: {total_count} entries for all regions")

def start_processing(selected_folder, region, summary_list_path, log_widget):
    def log_callback(msg):
        log_widget.config(state='normal')
        log_widget.insert(tk.END, msg + '\n')
        log_widget.see(tk.END)
        log_widget.config(state='disabled')
        log_widget.update_idletasks()
    # Pass export folder to processing
    export_folder = export_folder_var.get() if 'export_folder_var' in locals() else selected_folder
    threading.Thread(
        target=process_excels, 
        args=(selected_folder, None, summary_list_path, log_callback, export_folder), 
        daemon=True
    ).start()

def run_ui():


    root = tk.Tk()
    root.title("Excel Counter & Summary Updater - Enhanced Version")
    root.geometry("700x500")

    folder_var = tk.StringVar()
    folder_var.set(os.path.dirname(os.path.abspath(__file__)))

    def select_folder():
        folder_selected = filedialog.askdirectory()
        if folder_selected:
            folder_var.set(folder_selected)

    # Folder selection
    tk.Label(root, text="Select Folder:", font=('Arial', 10, 'bold')).pack(pady=5)
    folder_frame = tk.Frame(root)
    folder_frame.pack(pady=5)
    tk.Entry(folder_frame, textvariable=folder_var, width=60).pack(side=tk.LEFT)
    tk.Button(folder_frame, text="Browse", command=select_folder).pack(side=tk.LEFT, padx=5)

    # Summary list selection and export folder selection
    summary_list_path_var = tk.StringVar()
    summary_list_path_var.set("")
    export_folder_var = tk.StringVar()
    export_folder_var.set(os.path.dirname(os.path.abspath(__file__)))

    def select_summary_list():
        file_selected = filedialog.askopenfilename(
            title="Select existing summary_list.xlsx (optional)", 
            filetypes=[("Excel files", "*.xlsx")]
        )
        if file_selected:
            summary_list_path_var.set(file_selected)

    def select_export_folder():
        folder_selected = filedialog.askdirectory(title="Select folder to export summary_list.xlsx")
        if folder_selected:
            export_folder_var.set(folder_selected)

    tk.Label(root, text="Summary List (optional - will create if not selected):", font=('Arial', 10, 'bold')).pack(pady=(10,5))
    summary_frame = tk.Frame(root)
    summary_frame.pack(pady=5)
    tk.Button(summary_frame, text="Select summary_list.xlsx", command=select_summary_list).pack(side=tk.LEFT)
    summary_entry = tk.Entry(summary_frame, textvariable=summary_list_path_var, width=50, state='readonly')
    summary_entry.pack(side=tk.LEFT, padx=5)
    tk.Button(summary_frame, text="Select Export Folder", command=select_export_folder).pack(side=tk.LEFT, padx=5)
    export_entry = tk.Entry(summary_frame, textvariable=export_folder_var, width=40, state='readonly')
    export_entry.pack(side=tk.LEFT, padx=5)
    def clear_summary():
        summary_list_path_var.set("")
    tk.Button(summary_frame, text="Clear", command=clear_summary).pack(side=tk.LEFT, padx=5)

    # Log widget
    tk.Label(root, text="Processing Log:", font=('Arial', 10, 'bold')).pack(pady=(10,5))
    log_widget = scrolledtext.ScrolledText(root, width=80, height=15, state='disabled')
    log_widget.pack(pady=5, padx=10, fill='both', expand=True)

    def on_start():
        if not folder_var.get():
            messagebox.showerror("Error", "Please select a folder first!")
            return
        
        log_widget.config(state='normal')
        log_widget.delete(1.0, tk.END)
        log_widget.config(state='disabled')
        
        start_processing(folder_var.get(), None, summary_list_path_var.get(), log_widget)


    # Start button
    start_btn = tk.Button(root, text="Start Processing", command=on_start, 
                         width=20, height=2, bg='#4CAF50', fg='white', 
                         font=('Arial', 12, 'bold'))
    start_btn.pack(pady=10)

    # --- Generate summary_list.xlsx from file_amounts.txt ---
    def generate_summary_list_xlsx():
        folder = folder_var.get()
        file_amounts_path = os.path.join(folder, 'file_amounts.txt')
        summary_xlsx_path = os.path.join(folder, 'summary_list.xlsx')
        log_widget.config(state='normal')
        log_widget.insert(tk.END, f"Generating summary_list.xlsx from file_amounts.txt...\n")
        log_widget.see(tk.END)
        log_widget.config(state='disabled')
        log_widget.update_idletasks()
        if not os.path.exists(file_amounts_path):
            log_widget.config(state='normal')
            log_widget.insert(tk.END, f"file_amounts.txt not found in {folder}\n")
            log_widget.see(tk.END)
            log_widget.config(state='disabled')
            return
        # Parse file_amounts.txt
        region_date_counts = {}
        with open(file_amounts_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                if line.startswith('[') and line.endswith(']'):
                    continue
                # Format: YYYY-MM-DD - REGION - COUNT
                try:
                    date_part, region_part, count_part = [x.strip() for x in line.split('-', 2)]
                    region = region_part
                    count = int(count_part)
                    date = date_part
                    if region not in region_date_counts:
                        region_date_counts[region] = {}
                    if date not in region_date_counts[region]:
                        region_date_counts[region][date] = 0
                    region_date_counts[region][date] += count
                except Exception as e:
                    continue
        # Collect all unique dates and convert to datetime.date
        import datetime
        all_dates = set()
        for region in region_date_counts:
            for d in region_date_counts[region]:
                try:
                    all_dates.add(datetime.datetime.strptime(d, '%Y-%m-%d').date())
                except:
                    pass
        all_dates = sorted(all_dates)
        # Write to summary_list.xlsx using create_or_update_summary_list
        for region in region_date_counts:
            # Convert date keys to datetime.date
            date_counts = {}
            for d, v in region_date_counts[region].items():
                try:
                    d_obj = datetime.datetime.strptime(d, '%Y-%m-%d').date()
                    date_counts[d_obj] = v
                except:
                    continue
            create_or_update_summary_list(summary_xlsx_path, region, date_counts, all_dates, log_widget)
        log_widget.config(state='normal')
        log_widget.insert(tk.END, f"summary_list.xlsx generated at {summary_xlsx_path}\n")
        log_widget.see(tk.END)
        log_widget.config(state='disabled')
        log_widget.update_idletasks()

    # Generate summary_list.xlsx button
    summary_xlsx_btn = tk.Button(root, text="Generate summary_list.xlsx", command=generate_summary_list_xlsx,
                                width=25, height=2, bg='#2196F3', fg='white', font=('Arial', 12, 'bold'))
    summary_xlsx_btn.pack(pady=5)

    # Add some instructions
    instructions = tk.Label(root, 
                           text="ENHANCED VERSION: Better error handling, detailed logging, and improved counting logic.\n"
                                "Select folder with Excel files, optionally select existing summary file.\n"
                                "The tool will show detailed progress and identify any counting issues.\n"
                                "Region is now automatically detected from column E in each file.",
                           font=('Arial', 9), fg='gray', wraplength=650)
    instructions.pack(pady=5)

    root.mainloop()

if __name__ == "__main__":
    run_ui()